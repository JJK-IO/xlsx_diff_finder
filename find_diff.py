import sys

from openpyxl import load_workbook

print("Loading %s..." % sys.argv[1])
wb1 = load_workbook(sys.argv[1], data_only=True)
ws1 = wb1.active

print("Loading %s..." % sys.argv[2])
wb2 = load_workbook(sys.argv[2], data_only=True)
ws2 = wb2.active

max_rows = max(ws1.max_row, ws2.max_row)
print("max row: %s" % max_rows)

# max_columns = max(ws1.max_column, ws2.max_column)
max_columns = 18
print("max columns: %s" % max_columns)

cache = {}

headers = []
color_headers = ['j_color', 'k_color', 'l_color', 'm_color', 'n_color']
color_start_index = 10

for column in range(1, max_columns + 1):
    if ws1.cell(row=1, column=column).value is not None:
        headers.append(ws1.cell(row=1, column=column).value)
headers.extend(color_headers)

for cur_row in range(2, max_rows + 1):
    for cur_column in range(0, len(headers) - 5):
        cache.setdefault(headers[cur_column], []).append(ws1.cell(row=cur_row, column=cur_column + 1).value)
    cur_color_index = color_start_index
    for cur_column in range(len(headers) - 5, len(headers)):
        cache.setdefault(headers[cur_column], []).append(ws1.cell(row=cur_row, column=cur_color_index).fill.start_color.index)
        cur_color_index += 1

for cur_row in range(2, max_rows + 1):
    if ws2.cell(row=cur_row, column=4).value is not None:
        for cur_index in range(0, len(cache[ws2.cell(row=1, column=4).value])):
            if ws2.cell(row=cur_row, column=4).value == cache[ws2.cell(row=1, column=4).value][cur_index]:
                # Compare values from ws2 ws to c1 dictionary
                for cur_column in range(0, len(headers) - 5):
                    if ws2.cell(row=cur_row, column=cur_column + 1).value != \
                            cache[ws2.cell(row=1, column=cur_column + 1).value][cur_index] and \
                                    ws2.cell(row=cur_row, column=cur_column + 1).value is None:
                        print("WS1\t[%s]" % ws2.cell(row=cur_row, column=4).value)
                        print("<%s" % cache[ws2.cell(row=1, column=cur_column + 1).value][cur_index])
                        print(">%s" % ws2.cell(row=cur_row, column=cur_column + 1).value)

                cur_color_index = color_start_index
                cur_color_header_index = 0
                for cur_column in range(len(headers) - 5, len(headers)):
                    try:
                        if ws2.cell(row=cur_row, column=cur_color_index).fill.start_color.index != \
                                cache[color_headers[cur_color_header_index]][cur_row]:
                            print("WS1\t[%s]" % ws2.cell(row=cur_row, column=4).value)
                            print("%s" % color_headers[cur_color_header_index])
                            print("<%s" % cache[color_headers[cur_color_header_index]][cur_row])
                            print(">%s" % ws2.cell(row=cur_row, column=cur_color_index).fill.start_color.index)
                    except IndexError:
                        pass
                    cur_color_index += 1
                    cur_color_header_index += 1

